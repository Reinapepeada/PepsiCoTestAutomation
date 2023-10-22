#importo librerias de excel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os 

# creo una funcion que reciba una lista de listas y cree un excel con esos datos
def createExcel(contadorTotalList,equiposFallandolist,tubosdesconectados):

    # creo un libro de trabajo
    wb = Workbook()
    # creo una hoja de trabajo
    ws = wb.active
    # para cada lista voy a tener una hoja

    # creo una tabla con los datos de la lista
    table = Table(displayName="Table1", ref="A1:C"+str(len(contadorTotalList)))
    # creo un estilo para la tabla
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    # aplico el estilo a la tabla
    table.tableStyleInfo = style
    # agrego la tabla a la hoja
    ws.add_table(table)
    # recorro la lista y la agrego a la hoja
    for i in range(len(contadorTotalList)):
        ws.cell(row=i+1, column=1).value = contadorTotalList[i][0]
        ws.cell(row=i+1, column=2).value = contadorTotalList[i][1]
        ws.cell(row=i+1, column=3).value = contadorTotalList[i][2]

    # creo un estilo para las celdas
    font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

    # recorro la hoja y le aplico el estilo a las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    
    # creo otra hoja
    ws2 = wb.create_sheet("TC03MPC02DP")
    # creo una tabla con los datos de la lista
    table = Table(displayName="Table2", ref="A1:C"+str(len(equiposFallandolist)))
    # creo un estilo para la tabla
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    # aplico el estilo a la tabla
    table.tableStyleInfo = style
    # agrego la tabla a la hoja
    ws2.add_table(table)
    # recorro la lista y la agrego a la hoja
    for i in range(len(equiposFallandolist)):
        ws2.cell(row=i+1, column=1).value = equiposFallandolist[i][0]
        ws2.cell(row=i+1, column=2).value = equiposFallandolist[i][1]
        ws2.cell(row=i+1, column=3).value = equiposFallandolist[i][2]
    
    # creo un estilo para las celdas
    font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

    # recorro la hoja y le aplico el estilo a las celdas
    for row in ws2.iter_rows():
        for cell in row:
            cell.font = font


    # creo otra hoja

    ws3 = wb.create_sheet("TC03MPC02DP")

    # creo una tabla con los datos de la lista

    table = Table(displayName="Table3", ref="A1:C"+str(len(tubosdesconectados)))
    
    # creo un estilo para la tabla

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    
    # aplico el estilo a la tabla

    table.tableStyleInfo = style

    # agrego la tabla a la hoja

    ws3.add_table(table)

    # recorro la lista y la agrego a la hoja

    for i in range(len(tubosdesconectados)):
        ws3.cell(row=i+1, column=1).value = tubosdesconectados[i][0]
        ws3.cell(row=i+1, column=2).value = tubosdesconectados[i][1]
        ws3.cell(row=i+1, column=3).value = tubosdesconectados[i][2]
    
    # creo un estilo para las celdas

    font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

    # recorro la hoja y le aplico el estilo a las celdas

    for row in ws3.iter_rows():
        for cell in row:
            cell.font = font
    
    # buscar el path de la carpeta donde se encuentra el archivo

    path = os.path.dirname(os.path.abspath(__file__))

    # creo el path del archivo

    path = path + "\\TC03MPC02DPEfficien.xlsx"

    #no me crea el excel

    wb.save(path)



# creo el main
# creo una lista de listas
# cada lista interna es una fila de la tabla
# cada elemento de la lista interna es una columna de la tabla
lista = [["Date", "Efficiency", "Capacity"], ["2020-01-01", 0.9, 0.8], ["2020-01-02", 0.8, 0.7], ["2020-01-03", 0.7, 0.6], ["2020-01-04", 0.6, 0.5], ["2020-01-05", 0.5, 0.4], ["2020-01-06", 0.4, 0.3], ["2020-01-07", 0.3, 0.2], ["2020-01-08", 0.2, 0.1], ["2020-01-09", 0.1, 0.0]]
# llamo a la funcion
# le paso la lista
createExcel(lista,lista,lista) 

