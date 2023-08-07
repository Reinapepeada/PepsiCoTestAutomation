# import openpyxl

# def crear_archivo_excel(nombre_archivo, hojas):
    
#     libro = openpyxl.Workbook()
    
#     for nombre_hoja, datos_hoja in hojas.items():
#         hoja = libro.create_sheet(nombre_hoja)
#         for fila in datos_hoja:
#             hoja.append(fila)
    
#     libro.save(nombre_archivo)


# # Ejemplo de uso
# if __name__ == "__main__":
#     datos_hoja1 = [
#         ["Nombre", "Edad", "Profesión"],
#         ["Juan", 30, "Ingeniero"],
#         ["María", 25, "Médico"],
#         ["Pedro", 28, "Abogado"],
#     ]

#     datos_hoja2 = [
#         ["Producto", "Precio"],
#         ["Libro", 15.99],
#         ["Computadora", 899.99],
#         ["Teléfono", 399.99],
#     ]

#     hojas = {
#         "Personas": datos_hoja1,
#         "Productos": datos_hoja2,
#     }

#     crear_archivo_excel("archivo_excel.xlsx", hojas)

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

